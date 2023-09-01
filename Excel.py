# 1. Imported Modules
# -------------------
import openpyxl

def ImportGameExcel(gameNumber, file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    list1 = []
    number = None
    for row in ws.rows:
        if row[2].value == gameNumber:
            for cell in row:
                if cell.row not in list1:
                    list1.append((cell.row))
                    number = list1[0]
    if number is not None:
        start_row = number-3
        max_row1 = start_row + 13
        list2 = []
        list3 = []
        for i in range(start_row, max_row1):
            
            cell_obj1 = ws.cell(row = i, column = 2)
            cell_obj2 = ws.cell(row = i, column = 3)
            if cell_obj1.value == None:
                break
            list2.append(cell_obj1.value)
            list3.append(cell_obj2.value)
        res = {}
        for key in list2:
            for value in list3:
                res[key] = value
                list3.remove(value)
                break
        whiteMoves = []
        blackMoves = []
        for i in range(max_row1, ws.max_row):
            cell_obj1 = ws.cell(row = i, column = 3)
            cell_obj2 = ws.cell(row = i, column = 4)
            if cell_obj1.value == None or cell_obj2.value == None:
                break
            whiteMoves.append(cell_obj1.value)
            blackMoves.append(cell_obj2.value)
        allMoves = [None]*(len(whiteMoves)+len(blackMoves))
        allMoves[::2] = whiteMoves
        allMoves[1::2] = blackMoves
        if 'White' in allMoves:
            allMoves.remove('White')
        if 'Black' in allMoves:
            allMoves.remove('Black')
        list4 = []
        list4.append(res)
        list4.append(allMoves)
        return list4
    else:
        print('Game not found')
        
def ExportGameExcel(game, file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    gameMeta = game[0]
    gameMoves = game[1]
    gameNumber = gameMeta['Round']
    gameNumbers = [row[2].value for row in ws.rows if row[2].value is not None]
    
    if gameNumber in gameNumbers:
        print(f"Round '{gameNumber}' already exists in the file")
    else:
        start_row = ws.max_row + 2
        for i, (key, value) in enumerate(gameMeta.items()):
            ws.cell(row=start_row + i, column=2).value = key
            ws.cell(row=start_row + i, column=3).value = value

        # Determine the row number to start writing at (one row after the last row with data)
        start_row = ws.max_row + 2
        ws.cell(row=start_row, column=2).value = 'Turn'
        ws.cell(row=start_row, column=3).value = 'white'
        ws.cell(row=start_row, column=4).value = 'Black'
        start_row = ws.max_row + 1
        # Loop through the list of values and write them to the Excel file
        for i in range(0, len(gameMoves), 2):
            ws.cell(row=start_row, column=2).value = i//2 + 1
            ws.cell(row=start_row, column=3).value = gameMoves[i]
            ws.cell(row=start_row, column=4).value = gameMoves[i+1] if i+1 < len(gameMoves) else ''
            # Increment the row number for the next iteration
            start_row += 1
        # Save the changes to the Excel file
        wb.save(file_name)